# -*- coding: utf-8 -*-
"""초등영어 단어가 된다 1·2권: 문장 데이터를 모아 최종 엑셀로 만든다.

`ex_*.py`의 DATA(dict: english -> example)를 합쳐 추출 스켈레톤에 채운다.
이 교재는 동의어·반의어를 넣지 않는다(초등 기초어라 억지로 넣으면 학습이 안 됨).

검사:
  - 표제어가 문장에 **낱말 단위로** 들어 있는가
    (경계 검사 없이 하면 ear가 hear를, I가 This를 잡는다)
  - 문장 길이 3~7단어
  - 문장 중복
  - 저학년이 모를 단어를 쓰지 않았는가 — 교재 전체 표제어 + 아래 기초어만 허용
"""
import glob
import importlib.util
import os
import re
import sys
import collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# 초1~3이 이미 아는 것으로 보는 기초어(교과서 기본 문형·동작·성질에 쓰이는 말).
# 여기에 없고 교재 표제어도 아니면 검사에서 걸린다 — 저학년에 어려운 말을 막는 장치.
BASIC = set("""
a an the this that these those is are am be
i you he she it we they me him her us them
my your his its our their
have has like likes love loves know knows go goes come comes
eat eats drink drinks see sees play plays want wants
do does can and or but not no yes very too so
in on at to of for with from up down out by
here there now today very much many some any all
one two three four five six seven eight nine ten
big small long short good bad new old hot cold warm cool
is not isn't don't doesn't can't it's i'm
say says open opens close closes read reads write writes make makes
ride rides sleep sleeps work works ring rings need needs cut cuts
hold holds wave waves wash washes sing sings jump jumps run runs
walk walks sit sits stand stands swim swims dance dances give gives
put puts take takes look looks listen listens help helps
sweet kind strong tall clean soft heavy high tail best pretty every
today tonight tomorrow lunch dinner breakfast song paper wall
sky sun grass hat pool wet dry happy sad tall short high low
live lives made deep was were last well side please tea coffee
dollar dollars third fourth fifth next stop line street
seoul busan
age tell tells will soon pen lost leaves box thin almost lot ideas
smile beats wings turn turns idea
cannot must wide did done job
teeth feet legs arms hands eyes ears fingers toes friends brothers sisters
doctor nurse car bus rope boat word hour minute sticker game fun loud
after before together out again around only also just about
fly flies tv math music meet meets wait waits gold sharp bright dark
many our her his years days months tree flower toy ball book bag
""".split())


def known(token, allowed):
    """복수형·3인칭·진행형은 원형이 허용되면 함께 허용한다(eyes -> eye)."""
    if token in allowed:
        return True
    for cut, add in ((1, ""), (2, ""), (3, ""), (2, "y"), (3, "e")):
        if len(token) > cut and token[:-cut] + add in allowed:
            if token.endswith(("s", "es", "ing", "ed", "ies")):
                return True
    return False


def load_data():
    data, where = {}, {}
    for path in sorted(glob.glob(os.path.join(HERE, "ex_*.py"))):
        name = os.path.basename(path)[:-3]
        spec = importlib.util.spec_from_file_location(name, path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules[name] = mod
        spec.loader.exec_module(mod)
        for k, v in getattr(mod, "DATA", {}).items():
            if k in data:
                print("  !! 중복 정의:", k, "(", where[k], "/", name, ")")
            data[k], where[k] = v, name
    return data


def primary(eng):
    return re.split(r"[\[(]", eng)[0].strip()


def build(src, out, data, book_words):
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2))
    # 교재에 실린 단어는 전부 허용(다른 DAY 단어도 문장에 쓸 수 있게).
    # 1·2권이 같은 단어를 일부 공유하므로 두 권을 합쳐서 본다.
    allowed = set(BASIC)
    for w in book_words:
        for t in re.findall(r"[A-Za-z']+", w.lower()):
            allowed.add(t)

    filled, noword, toolong, hard = 0, [], [], []
    for r in rows:
        eng = r[3].value
        ex = data.get(eng)
        if not ex:
            continue
        r[7].value = ex
        filled += 1
        p = primary(eng)
        if not re.search(r"(?<![A-Za-z])" + re.escape(p) + r"(?:s|es|ies|ing|ed)?(?![A-Za-z])",
                         ex, re.I):
            noword.append((eng, ex))
        n = len(ex.split())
        if not 3 <= n <= 7:
            toolong.append((eng, n, ex))
        # 고유명사(문장 중간의 대문자 시작)는 검사에서 뺀다
        body = re.sub(r"(?<!^)(?<![.!?] )[A-Z][a-z]+", "", ex)
        for t in re.findall(r"[A-Za-z']+", body.lower()):
            if not known(t, allowed):
                hard.append((eng, t, ex))
    return wb, rows, filled, noword, toolong, hard


def main():
    data = load_data()
    ok = True
    book_words = set()
    for n in (1, 2, 3, 4):
        ws = openpyxl.load_workbook(os.path.join(HERE, "cho%d.xlsx" % n),
                                    read_only=True).active
        for r in list(ws.iter_rows(min_row=2, values_only=True)):
            book_words.add(r[3])
    for n in (1, 2, 3, 4):
        src = os.path.join(HERE, "cho%d.xlsx" % n)
        out = os.path.join(HERE, "초등영어 단어가 된다_%d.xlsx" % n)
        wb, rows, filled, noword, toolong, hard = build(src, out, data, book_words)
        total = len(rows)
        print("%d권: %d / %d 문장 (%.0f%%)" % (n, filled, total, filled / total * 100))
        for label, items in (("표제어가 문장에 없음", noword),
                             ("길이가 3~7단어를 벗어남", toolong),
                             ("교재에 없는 어려운 낱말", hard)):
            if items:
                ok = False
                print("  !! %s: %d건" % (label, len(items)))
                for it in items[:14]:
                    print("     ", it)
        if filled == total:
            wb.save(out)
            print("  저장:", out)
    dup = [(x, c) for x, c in collections.Counter(data.values()).items() if c > 1]
    if dup:
        ok = False
        print("!! 문장 중복:")
        for x, c in dup:
            print("   (%d회) %s" % (c, x))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
