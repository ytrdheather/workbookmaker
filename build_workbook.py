# -*- coding: utf-8 -*-
"""
어휘 워크북 자동 생성기
- 입력: 엑셀 (book_name, unit_name, word_no, english, meaning, synonyms, antonyms, example)
- 출력: 유닛(DAY)별로 규칙에 맞는 워크북 HTML -> Chrome 헤드리스로 PDF

유닛 페이지 규칙:
  1. 단어표 (english / 의미 / 예문 / 반의어 / 동의어)   * 품사 열 없음
  2. 4회 쓰기 (english 3번 + 뜻 1번), 순서 랜덤
  [ N-2 유닛 복습 시험 ]  (존재할 때만, 매번 다른 버전)
  [ N-1 유닛 복습 시험 ]  (존재할 때만, 매번 다른 버전)
  5. 예문 빈칸 (의미 + 빈칸 예문), 순서 랜덤
  6. 반의어 고르기 (4지선다)
  7. 동의어 고르기 (4지선다)
"""
import openpyxl
import math
import random
import html
import re
import os
import sys
import subprocess
import argparse

# --layout v2 (Bricks 4800 전용). 기본 v1은 기존 교재와 동일한 지면을 그대로 유지한다.
#   v2: 로고를 지면별로 배치(암기노트·쓰기는 머리말 오른쪽), 쓰기 연습을 22행+나머지 2장으로 분할,
#       하단 로고 자리를 본문에서 비워 겹치지 않게 함.
LAYOUT2 = False

# --fit-choice: 동의/반의 합친 지면이 조금 넘칠 때 여백·글자를 줄여 한 장에 밀어 넣는다.
# 기본 꺼짐(끄면 넘치는 만큼 다음 장으로 — 기존 교재 지면 그대로).
FIT_CHOICE = False

# --logo-top: v2의 로고 배치만 떼어 쓴다. 암기 노트·쓰기 연습은 로고를 머리말 오른쪽
# (학습한 날짜 뒤)으로 올리고, 나머지 지면은 우하단에 지면별로 넣는다. 하단이 비므로
# 그 두 지면은 세로 예산을 늘려 잡는다. v2와 달리 쓰기 연습을 분할하지 않는다.
LOGO_TOP = False


def logo_on_top():
    """암기 노트·쓰기 연습에서 로고를 머리말로 올리는가."""
    return LAYOUT2 or LOGO_TOP


# --write-split N: 쓰기 연습을 N행씩 나눈다(0이면 분할 없이 한 장).
# 단어가 많은 교재에서 쓰기 칸이 너무 눌리는 것을 푸는 용도.
WRITE_SPLIT = 0

# --split-wordlist N: 단어 목록을 N개씩 나눈다(0이면 한 장).
#   판매용 제본은 93% 축소를 거치므로 30단어를 한 장에 우겨넣으면 인쇄물이 7.4pt가 된다.
WORDLIST_SPLIT = 0
# --memo-split N: 단어 암기 노트를 N행씩 나눈다(0이면 한 장).
# 30단어를 한 장에 넣으면 행이 18px까지 눌려 쓸 수 없어지는 것을 푸는 용도.
MEMO_SPLIT = 0

# 쓰기/암기를 나눈 지면은 남는 높이를 행에 돌려준다(안 그러면 아래가 크게 빔).
# 뜻이 2줄로 접히는 행이 있어 그만큼 여유를 두고 상한을 잡았다(실측값).
SPLIT_ROW_MAX_WRITE = 44
SPLIT_ROW_MAX_MEMO = 44
SPLIT_BODY_PX_WRITE = 860
SPLIT_BODY_PX_MEMO = 700

# --balance-choice: 고르기가 두 장 이상으로 나뉠 때 앞 장을 꽉 채우지 않고 고르게 나눈다.
# (24 + 나머지 2개 같은 지면을 막는다.)
BALANCE_CHOICE = False

# --fit-wordlist: 단어 목록이 한 장을 넘길 때 여백·글자를 더 줄여 한 유닛을 한 장에 담는다.
FIT_WORDLIST = False

# --flat-review: 누적 복습 시험의 문제(단어/뜻)가 두 줄로 접히지 않도록
# 문제 칸을 넓히고, 그래도 넘치는 만큼 글자를 줄인다.
FLAT_REVIEW = False

# --fill-review: 누적 복습 시험에서 아래로 남는 여백을 행 간격으로 돌려 지면을 채운다.
# 표 위 여백 70pt, 하단 로고 시작 769pt 기준으로 쓸 수 있는 높이가 약 932px.
FILL_REVIEW = False
REVIEW_BODY_PX = 932


def _text_w(s, fs):
    """문자열의 대략적인 렌더 폭(px). 한글·전각은 1em, 그 외는 0.55em으로 잡는다."""
    return sum((1.0 if ord(c) > 0x1100 else 0.55) for c in s) * fs


# 단어 목록 표에 쓸 수 있는 높이(표머리 top ~ 로고 top). 실측하면 v1은 913px,
# v2/logo-top 교재는 909px로 4px 빠듯하다. 둘 다 안전하도록 900으로 잡는다
# (행당 0.4px 차이라 눈에 안 보인다). 표 머리 행 34px도 실측.
# 열별 글자 폭은 실측 열 폭에서 좌우 여백 7px씩 뺀 값.
WL_BODY_PX = 900
WL_THEAD_PX = 34
# 행 높이 계수. CSS line-height는 1.3이지만 셀 여백·테두리·줄바꿈 반올림이 더해져
# 실제로는 더 든다. 넘치지 않는 쪽으로 잡은 실측 보정값(_wl_pad 주석 참고).
WL_LINE_H = 1.45
WL_COL_AVAIL = {"english": 67, "meaning": 135, "example": 272,
                "antonyms": 65, "synonyms": 65}


def _wl_pad(words, fs, base=9, cap=24):
    """--split-wordlist 지면에서 남는 높이를 셀 여백으로 돌려준다.

    나눈 뒤 남는 높이는 교재마다 다르다(고난도는 뜻이 길어 16mm, 3900은 짧아 87mm).
    한 값으로 박으면 한쪽은 넘치고 한쪽은 텅 빈다. 그래서 지면마다 계산한다.

    글자폭 계수(한글 0.92 / 영문 0.41)와 행간 1.45는 네 교재 300장을 실측해 맞춘 값으로,
    **절대 과소평가하지 않도록** 골랐다(실측 대비 +10~+173px). 과소평가하면 여백을
    과하게 줘서 표가 로고를 덮는다 — 실제로 행간 1.3으로 잡았을 때 Bricks 4800이
    24px 모자라 1.4mm 겹쳤다. 남는 쪽으로 틀려야 안전하다.
    base 미만으로는 내리지 않는다."""
    rows = 0
    for w in words:
        n = 1
        for key, avail in WL_COL_AVAIL.items():
            t = w.get(key) or "-"
            px = sum((0.92 if ord(c) > 0x1100 else 0.41) for c in t) * fs
            n = max(n, math.ceil(px / avail) or 1)
        rows += n
    cnt = max(len(words), 1)
    used = WL_THEAD_PX + rows * fs * WL_LINE_H + cnt * (2 * base + 1)
    return int(max(base, min(cap, base + (WL_BODY_PX - used) / (2 * cnt))))

# ---------------------------------------------------------------- 데이터 로드
def load_words(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx = {name: header.index(name) for name in header}

    def col(r, *names):
        for n in names:
            if n in idx and idx[n] < len(r):
                v = r[idx[n]]
                return ("" if v is None else str(v)).strip()
        return ""

    books = {}   # book_name -> list of unit dicts (in order)
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        book = col(r, "book_name") or "Workbook"
        unit = col(r, "unit_name") or "UNIT"
        w = {
            "no": col(r, "word_no"),
            "english": col(r, "english"),
            "meaning": col(r, "meaning"),
            "synonyms": col(r, "synonyms", "synonym"),
            "antonyms": col(r, "antonyms", "antonym"),
            "example": col(r, "example"),
        }
        if not w["english"]:
            continue
        books.setdefault(book, {})
        books[book].setdefault(unit, [])
        books[book][unit].append(w)

    # dict 순서 유지 -> 유닛 리스트로
    result = {}
    for book, units in books.items():
        result[book] = [(uname, words) for uname, words in units.items()]
    return result


def split_multi(val):
    """'repair, mend' -> ['repair','mend']"""
    if not val:
        return []
    parts = re.split(r"[,/;]", val)
    return [p.strip() for p in parts if p.strip()]


# ---------------------------------------------------------------- HTML 조각
def esc(s):
    return html.escape(s or "")


def blank_out(example, english):
    """예문에서 대상 단어를 빈칸으로. 못 찾으면 문장 뒤에 빈칸 표시."""
    if not example:
        return "_________"
    pat = re.compile(re.escape(english), re.IGNORECASE)
    if pat.search(example):
        return pat.sub('<span class="blank"></span>', example, count=1)
    return esc(example) + ' ( <span class="blank"></span> )'


# ---------------------------------------------------------------- 페이지: 단어표
def page_wordlist(day_name, words):
    """단어 목록. WORDLIST_SPLIT이 있으면 N개씩 나눠 여러 장으로. -> 페이지 리스트"""
    total = max(len(words), 1)
    step = WORDLIST_SPLIT if WORDLIST_SPLIT else total
    return [_wordlist_page(day_name, words[s:s + step], ci)
            for ci, s in enumerate(range(0, total, max(step, 1)))]


def _wordlist_page(day_name, words, ci=0):
    rows = []
    for w in words:
        rows.append(f"""
      <tr>
        <td class="c-no">{esc(w['no'])}</td>
        <td class="c-eng">{esc(w['english'])}</td>
        <td class="c-mean">{esc(w['meaning'])}</td>
        <td class="c-ex">{esc(w['example'])}</td>
        <td class="c-ant">{esc(w['antonyms'] or '-')}</td>
        <td class="c-syn">{esc(w['synonyms'] or '-')}</td>
      </tr>""")
    # 단어 수가 많아도 한 페이지에 들어가도록 여백/글자크기 자동 축소
    n = max(len(words), 1)
    if WORDLIST_SPLIT:
        # 판매용 제본 대비. make_print_ready.py가 93% 축소하므로 인쇄물에서 9pt를
        # 넘기려면 축소 전 9.7pt(약 13px) 이상이어야 한다. 나눠서 행이 줄었으니 여유 있음.
        # 남는 높이는 지면마다 달라서 여백으로 돌려준다(_wl_pad 주석 참고).
        wlf = 13.5
        wlp = _wl_pad(words, wlf)
    elif LAYOUT2:   # 하단 로고 자리를 비워도 여유가 있어 오히려 키움
        wlp = 7 if n <= 20 else (6 if n <= 25 else 5)
        wlf = 12.5 if n <= 20 else (12 if n <= 25 else 11.5)
    elif FIT_WORDLIST and n > 25:
        # 예문·의미가 길어 마지막 한두 행이 다음 장으로 넘치던 교재용(중등 고난도 30단어).
        wlp, wlf = 2, 10.5
    else:
        wlp = 7 if n <= 20 else (5 if n <= 25 else 4)
        wlf = 12.5 if n <= 20 else (11.5 if n <= 25 else 11)
    return f"""
  <section class="page">
    {page_head(day_name, f"단어 목록{' (계속)' if ci else ''}", "Word List")}
    <table class="wl" style="--wlp:{wlp}px; --wlf:{wlf}px">
      <thead>
        <tr><th>No.</th><th>단어</th><th>의미</th><th>예문</th><th>반의어</th><th>동의어</th></tr>
      </thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    {page_foot()}
  </section>"""


# ---------------------------------------------------------------- 페이지: 암기 노트
def page_memorize(day_name, words):
    """딸기케이크식 4단 자가시험 노트(빈칸). 단어를 직접 쓰게 하고, 단어 수에 맞춰 행 높이 자동 조정.
    단어목록 바로 뒤에 배치해 '외우는 단계'를 강제한다. -> 페이지 리스트"""
    total = max(len(words), 1)
    step = MEMO_SPLIT if MEMO_SPLIT else total
    return [_memorize_page(day_name, min(step, total - s), s, ci)
            for ci, s in enumerate(range(0, total, max(step, 1)))]


def _memorize_page(day_name, n, offset, ci):
    n = max(n, 1)
    if LAYOUT2:
        # 로고를 머리말로 올려 하단까지 다 쓴다. 예산 안에서 행을 최대한 키우고,
        # 남는 높이에 맞춰 오답 정리 줄 수를 정한다(모자라면 오답 정리를 뺀다).
        row_h = max(24, min(34, 855 // n))
        oa_lines = max(0, min(6, (855 - row_h * n - 40) // 25))
    elif LOGO_TOP:
        # 로고가 머리말로 올라가 하단이 비므로 행 예산을 늘린다. 오답 정리 6줄은
        # 고정으로 두고(855에서 그 상자 높이 170만큼 뺀 685), 남는 만큼 행을 키운다.
        row_h = max(17, min(34, 685 // n))
        oa_lines = 6
    elif MEMO_SPLIT:
        # 나눠서 행이 적어진 만큼 행을 키워 지면을 채운다(오답 정리 6줄은 그대로).
        row_h = max(17, min(SPLIT_ROW_MAX_MEMO, SPLIT_BODY_PX_MEMO // n))
        oa_lines = 6
    else:
        # 하단 로고 여백을 확보하며 한 페이지에. 오답 정리 6줄만큼 행 예산을 줄임(600 -> 545).
        row_h = max(17, min(34, 545 // n))
        oa_lines = 6
    rows = []
    for i in range(offset + 1, offset + n + 1):
        rows.append(
            f'<tr><td class="mz-no">{i}</td>'
            f'<td class="mz-fold"></td><td class="mz-fold"></td><td class="mz-fold"></td>'
            f'<td class="mz-fold mz-mark">□</td></tr>')
    oa = "".join('<div class="mz-oa-line"></div>' for _ in range(oa_lines))
    oa_html = (f'<div class="mz-oa"><div class="mz-oa-t">⑤ 오답 정리 — 틀린 단어만 다시 '
               f'(단어 + 뜻)</div>{oa}</div>') if oa_lines else ""
    return f"""
  <section class="page{' page-hl' if logo_on_top() else ''}">
    {page_head(day_name, f"단어 암기 노트{' (계속)' if ci else ''}", "Memorize &amp; Self-Test", logo_top=True)}
    <p class="guide"><b>① 단어</b> — 단어 목록을 보고 영어를 옮겨 씁니다. &nbsp;
      <b>② 뜻</b> — 단어를 보고 뜻을 <b>외워서</b> 씁니다(모르면 다른 색 펜). &nbsp;
      <b>③ 스펠링</b> — 오른쪽 점선을 접어 단어를 가리고, 뜻만 보고 씁니다. &nbsp;
      <b>④ 모름 □</b> 체크 → 아래 오답 정리에 <b>모르는 것만</b> 다시.</p>
    <table class="mz" style="--mzh:{row_h}px">
      <colgroup><col style="width:26px"><col style="width:22%"><col><col style="width:27%"><col style="width:34px"></colgroup>
      <thead><tr>
        <th>No</th><th class="mz-fold">① 단어</th>
        <th class="mz-fold">② 뜻 쓰기<br><span class="mz-th-sub">단어 보고 · 외워서</span></th>
        <th class="mz-fold">③ 스펠링 쓰기<br><span class="mz-th-sub">단어 접어 가리고 · 뜻만 보고</span></th>
        <th class="mz-fold">모름</th>
      </tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    {oa_html}
  </section>"""


# ---------------------------------------------------------------- 페이지: 4회 쓰기
WRITE_SLOTS = 22   # 쓰기 연습 한 장에 넣는 행 수(넘치면 다음 장으로)


def page_writing(day_name, words, seed):
    """쓰기 연습. 행이 눌리지 않도록 한 장에 WRITE_SLOTS행까지만 넣고 나머지는 다음 장. -> 페이지 리스트"""
    ws = words[:]
    random.Random(seed).shuffle(ws)
    if WRITE_SPLIT:
        step = WRITE_SPLIT
    elif LAYOUT2:
        step = WRITE_SLOTS
    else:
        step = len(ws)   # 기본 v1은 분할 없이 한 장
    chunks = [ws[s:s + step] for s in range(0, len(ws), max(step, 1))] or [[]]
    pages = []
    for ci, chunk in enumerate(chunks):
        rows = []
        for j, w in enumerate(chunk, ci * step + 1):
            rows.append(f"""
      <tr>
        <td class="w-no">{j}</td>
        <td class="w-eng">{esc(w['english'])}</td>
        <td class="w-mean">{esc(w['meaning'])}</td>
        <td class="w-write"><span class="wl3"></span><span class="wl3"></span><span class="wl3"></span></td>
      </tr>""")
        n = max(len(chunk), 1)
        if logo_on_top():   # 로고를 머리말로 올렸으므로 하단까지 쓴다
            # 뜻이 길어 2줄로 접히는 행이 있어, 행이 많으면 최소 높이를 낮춰 한 장에 유지한다.
            row_h = max(23, min(42, 860 // n))
            mean_fs = 12.5 if n <= 22 else (11.5 if n <= 26 else 11)
        elif WRITE_SPLIT:
            # 나눠서 행이 적어진 만큼 행을 키워 지면을 채운다.
            row_h = max(21, min(SPLIT_ROW_MAX_WRITE, SPLIT_BODY_PX_WRITE // n))
            mean_fs = 12.5 if n <= 20 else (11.5 if n <= 25 else 11)
        else:
            # 뜻이 길어 2줄로 줄바꿈되는 행이 있어 여유를 더 둠(25단어에서 넘치던 문제)
            row_h = max(21, min(38, 660 // n))
            mean_fs = 12.5 if n <= 20 else (11.5 if n <= 25 else 11)
        cont = " (계속)" if ci else ""
        pages.append(f"""
  <section class="page{' page-hl' if logo_on_top() else ''}">
    {page_head(day_name, f"쓰기 연습{cont}", "Write &amp; Remember", logo_top=True)}
    <p class="guide">뜻과 단어를 확인하고, 오른쪽 줄에 <b>영어 단어를 3번</b> 따라 쓰세요. <b>따라 쓸 때는 반드시 큰 소리로 단어를 읽으세요.</b></p>
    <table class="wr" style="--wrh:{row_h}px; --wmf:{mean_fs}px">
      <thead>
        <tr><th class="wh-no">No.</th><th class="wh-eng">단어</th><th class="wh-mean">뜻</th><th>3번 쓰기</th></tr>
      </thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
  </section>""")
    return pages


# ---------------------------------------------------------------- 페이지: 예문 빈칸
def _fillblank_parts(words, seed, answer):
    """예문 빈칸 은행/문항 HTML 생성 -> (bank_html, items_html)."""
    ws = words[:]
    random.Random(seed).shuffle(ws)
    bank = [w["english"] for w in ws]
    random.Random(seed + 1).shuffle(bank)
    items = []
    for w in ws:
        sent = blank_out(w["example"], w["english"])
        ans = f'<span class="ans">{esc(w["english"])}</span>' if answer else ""
        items.append(f"""
      <li>
        <span class="fb-sent">{sent}</span> {ans}
        <span class="fb-mean">({esc(w['meaning'])})</span>
      </li>""")
    bank_html = "".join(f'<span class="chip">{esc(b)}</span>' for b in bank)
    return bank_html, "".join(items)


def page_fillblank(day_name, words, seed, answer=False):
    bank_html, items_html = _fillblank_parts(words, seed, answer)
    # 문항 수가 많아도 한 페이지에 들어가도록 간격/글자크기/빈칸폭 자동 축소
    n = max(len(words), 1)
    fbm = 17 if n <= 20 else (11 if n <= 25 else 7)
    fbl = 1.7 if n <= 20 else (1.5 if n <= 25 else 1.35)
    fbf = 13 if n <= 20 else (12 if n <= 25 else 11.5)
    fbb = 120 if n <= 20 else (100 if n <= 25 else 88)
    return f"""
  <section class="page">
    {page_head(day_name, "예문 빈칸 채우기", "Fill in the Blank")}
    <p class="guide">뜻을 참고하여 빈칸에 알맞은 단어를 <b>단어 은행</b>에서 골라 쓰세요.</p>
    <div class="bank">{bank_html}</div>
    <ol class="fb" style="--fbm:{fbm}px; --fbl:{fbl}; --fbf:{fbf}px; --fbb:{fbb}px">{items_html}</ol>
    {page_foot()}
  </section>"""


# ---------------------------------------------------------------- 페이지: 복습 시험
def page_review(cur_day, target_day_name, words, seed, answer=False):
    """이전 유닛 단어↔의미 시험 (섞어서). seed로 매번 다른 버전."""
    rnd = random.Random(seed)
    ws = words[:]
    rnd.shuffle(ws)
    items = []
    qtexts = []
    for i, w in enumerate(ws, 1):
        ask_meaning = rnd.random() < 0.5  # True: 단어 주고 뜻 쓰기 / False: 뜻 주고 단어 쓰기
        if ask_meaning:
            q = esc(w["english"])
            a = esc(w["meaning"])
            qcls = "rv-eng"
        else:
            q = esc(w["meaning"])
            a = esc(w["english"])
            qcls = "rv-mean"
        qtexts.append(w["english"] if ask_meaning else w["meaning"])
        ansline = f'<span class="ans">{a}</span>' if answer else ''
        items.append(f"""
      <tr>
        <td class="rv-no">{i}</td>
        <td class="{qcls}">{q}</td>
        <td class="rv-a">{ansline}</td>
      </tr>""")
    # 단어 수가 많아도 한 페이지에 들어가도록 행 여백/글자크기 자동 축소(30단어에서 넘치던 문제)
    n = max(len(ws), 1)
    if LAYOUT2:
        pad = 9 if n <= 20 else (8 if n <= 25 else 6)
        rv_fs = 13 if n <= 20 else (12.5 if n <= 25 else 12)
    else:
        pad = 9 if n <= 20 else (6 if n <= 25 else 4)
        rv_fs = 13 if n <= 20 else (12 if n <= 25 else 11.5)
    qw_css = ""
    if FLAT_REVIEW:
        # 문제 칸을 넓혀 두 줄로 접히는 것을 없앤다. 답 쓰는 칸은 최소 260px 남긴다.
        QMAX = 430
        widest = max((_text_w(t, rv_fs) for t in qtexts), default=0) + 14
        if widest > QMAX:      # 그래도 넘치면 넘치는 비율만큼 글자를 줄임
            rv_fs = max(9.5, round(rv_fs * QMAX / widest, 1))
            widest = max((_text_w(t, rv_fs) for t in qtexts), default=0) + 14
        qw_css = f"; --rvq:{max(180, min(QMAX, int(widest) + 1))}px"
    if FILL_REVIEW:
        # 행 높이 = 글자 줄높이(1.35em) + 위아래 여백 + 경계선 1px.
        pad = max(pad, min(12, int((REVIEW_BODY_PX / n - rv_fs * 1.35 - 1) / 2)))
    return f"""
  <section class="page">
    {page_head(cur_day, f"누적 복습 시험 &middot; {esc(target_day_name)}", "Review Test")}
    <p class="guide">단어는 뜻을, 뜻은 단어를 쓰세요.</p>
    <table class="rv" style="--rvp:{pad}px; --rvf:{rv_fs}px{qw_css}">
      <tbody>{''.join(items)}</tbody>
    </table>
    {page_foot()}
  </section>"""


# ---------------------------------------------------------------- 페이지: 반의어/동의어 고르기
def _choice_items(words, kind, rnd, answer):
    """해당 kind(값 있는 단어)로 4지선다 항목 리스트 생성.
    <li>(list-item)는 Chrome 인쇄에서 break-inside:avoid가 무시돼 문제/선택지가 갈라지므로
    일반 블록 div(.ch-item)로 만들고 번호를 직접 붙인다."""
    pool = [w for w in words if split_multi(w[kind])]
    all_english = [w["english"] for w in words]
    rnd.shuffle(pool)
    items = []
    for qno, w in enumerate(pool, 1):
        corrects = split_multi(w[kind])
        correct = corrects[0]
        # 오답: 같은 유닛의 다른 english (정답/문제어 제외)
        avoid = set([w["english"], correct] + corrects)
        distract_pool = [e for e in all_english if e not in avoid]
        rnd.shuffle(distract_pool)
        options = [correct] + distract_pool[:3]
        rnd.shuffle(options)
        letters = "①②③④"
        opt_html = "".join(
            f'<span class="opt {"opt-ans" if (answer and o==correct) else ""}">{letters[j]} {esc(o)}</span>'
            for j, o in enumerate(options)
        )
        items.append(f"""
      <div class="ch-item">
        <div class="ch-q"><span class="ch-no">{qno}.</span> <b>{esc(w['english'])}</b> <span class="ch-mean">({esc(w['meaning'])})</span></div>
        <div class="ch-opts">{opt_html}</div>
      </div>""")
    return items


CHOICE_SLOTS = 24     # 고르기 페이지 한 장에 들어가는 항목 수(섹션 제목도 1칸 차지)
CHOICE_FIT_MAX = 30   # 이 칸수까지는 여백/글자를 줄여서라도 한 장에. 넘으면 그냥 다음 장으로.


def _choice_density(slots):
    """필요한 칸수에 맞춘 여백/글자 축소값. 기본 밀도(<=CHOICE_SLOTS)면 빈 문자열."""
    if slots <= CHOICE_SLOTS:
        return ""
    if slots <= 27:
        return "--chm:2px; --chq:11px; --cho:10.5px"
    # 30칸은 여백이 거의 안 남아, 앞 지면 누적 반올림에 따라 항목 하나가 다음 장으로
    # 새는 일이 있었다(중등 기본 DAY 55). 항목 간격을 0으로 둬 한 항목분 여유를 남긴다.
    return "--chm:0px; --chq:10.5px; --cho:10px; --chop:0px"


def _choice_pages(day_name, title_ko, title_en, guide, sections, slots=CHOICE_SLOTS, style=""):
    """고르기 항목을 페이지 단위로 '직접' 나눈다.
    Chrome이 인쇄에서 break-inside:avoid를 자주 무시해 문제/선택지가 갈라지므로,
    넘침을 브라우저에 맡기지 않고 여기서 항목 수를 세어 페이지를 끊는다.
    sections: [(섹션제목 or None, [항목 html, ...]), ...]"""
    pages, cur, used = [], [], 0

    def flush():
        nonlocal cur, used
        if not cur:
            return
        body = f'<div class="ch-body" style="{style}">{"".join(cur)}</div>' if style else "".join(cur)
        pages.append(f"""
  <section class="page">
    {page_head(day_name, title_ko, title_en)}
    {guide}
    {body}
    {page_foot()}
  </section>""")
        cur, used = [], 0

    for label, items in sections:
        if not items:
            items = ['<div class="empty">출제할 단어가 부족합니다.</div>']
        head_html = f'<div class="ch-sec">■ {label}</div>' if label else ""
        need_head = bool(label)
        opened = False          # 이 페이지에 ch-list를 열었는지
        for it in items:
            cost = 1 + (1 if need_head else 0)
            if used + cost > slots and cur:
                if opened:
                    cur.append("</div>")
                    opened = False
                flush()
                need_head = bool(label)
            if need_head:
                cur.append(head_html); used += 1; need_head = False
            if not opened:
                cur.append('<div class="ch-list">'); opened = True
            cur.append(it); used += 1
        if opened:
            cur.append("</div>")
    flush()
    return pages


def page_choice(day_name, words, kind, seed, answer=False):
    """kind: 'antonyms' or 'synonyms'. 해당 값 있는 단어만 출제. -> 페이지 리스트"""
    label = "반의어" if kind == "antonyms" else "동의어"
    en_label = "Antonym" if kind == "antonyms" else "Synonym"
    items = _choice_items(words, kind, random.Random(seed), answer)
    guide = f'<p class="guide">각 단어의 <b>{label}</b>를 보기에서 고르세요.</p>'
    return _choice_pages(day_name, f"{label} 고르기", f"Choose the {en_label}",
                         guide, [(None, items)])


def page_choice_merged(day_name, words, seed, answer=False):
    """동의어+반의어를 한 페이지에(넘치면 자동으로 다음 장). -> 페이지 리스트"""
    sitems = _choice_items(words, "synonyms", random.Random(seed), answer)
    aitems = _choice_items(words, "antonyms", random.Random(seed + 50), answer)
    guide = ('<p class="guide">각 단어의 <b>동의어(비슷한 말)</b> 또는 '
             '<b>반의어(반대말)</b>를 보기에서 고르세요.</p>')
    # 필요한 칸수(항목 + 섹션 제목 2개). 조금만 줄이면 한 장에 들어가는 경우는 줄여서 한 장에.
    need = max(len(sitems), 1) + max(len(aitems), 1) + 2
    slots, style = CHOICE_SLOTS, ""
    if FIT_CHOICE and CHOICE_SLOTS < need <= CHOICE_FIT_MAX:
        slots, style = need, _choice_density(need)
    elif BALANCE_CHOICE and need > CHOICE_SLOTS:
        # 앞 장을 24칸까지 채우면 뒷장에 2~3문항만 남는 지면이 생긴다. 장수는 그대로 두고
        # 고르게 나눈다(+1은 다음 장에서 섹션 제목이 다시 붙는 몫).
        npages = math.ceil(need / CHOICE_SLOTS)
        slots = math.ceil(need / npages) + 1
    return _choice_pages(day_name, "동의어 &middot; 반의어 고르기", "Choose Synonym / Antonym",
                         guide, [("동의어 고르기", sitems), ("반의어 고르기", aitems)],
                         slots=slots, style=style)


def page_practice_merged(day_name, words, seed, answer=False):
    """예문 빈칸 + 동의어 + 반의어를 한 페이지에. 단어 수가 아주 적은 교재(예: Bricks 300)용."""
    bank_html, fb_items = _fillblank_parts(words, seed + 4, answer)
    sitems = _choice_items(words, "synonyms", random.Random(seed + 5), answer)
    aitems = _choice_items(words, "antonyms", random.Random(seed + 6), answer)
    empty = '<div class="empty">출제할 단어가 부족합니다.</div>'
    syn_html = "".join(sitems) if sitems else empty
    ant_html = "".join(aitems) if aitems else empty
    return f"""
  <section class="page">
    {page_head(day_name, "종합 연습", "Practice")}
    <div class="ch-sec">■ 예문 빈칸 채우기 <span class="ch-hint">— 단어 은행에서 골라 쓰세요.</span></div>
    <div class="bank">{bank_html}</div>
    <ol class="fb">{fb_items}</ol>
    <div class="ch-sec">■ 동의어 고르기</div>
    <div class="ch-list">{syn_html}</div>
    <div class="ch-sec">■ 반의어 고르기</div>
    <div class="ch-list">{ant_html}</div>
    {page_foot()}
  </section>"""


# ---------------------------------------------------------------- 공통 헤더 / 로고
def page_head(day_name, title_ko, title_en, logo_top=False):
    """logo_top=True면 로고를 머리말 오른쪽(학습한 날짜 뒤)에 넣는다.
    쓰기 칸이 페이지 하단까지 꽉 차는 지면(암기 노트·쓰기 연습)에서 하단 로고와 겹치지 않게 하기 위함."""
    top = logo_top and logo_on_top()
    lg = f'<img class="ph-logo" src="{logo_datauri()}" alt="logo">' if top else ""
    return f"""
    <div class="phead{' phead-lg' if top else ''}">
      <div class="ph-day">{esc(day_name)}</div>
      <div class="ph-title"><span class="pt-ko">{title_ko}</span><span class="pt-en">{title_en}</span></div>
      <div class="ph-name">학습한 날짜 : <span class="nline"></span>{lg}</div>
    </div>"""


def page_foot():
    """지면 우하단 로고. 기본 v1은 body에 fixed 로고 하나를 두므로 지면별로는 넣지 않는다."""
    if not logo_on_top():
        return ""
    return f'<img class="pagelogo" src="{logo_datauri()}" alt="logo">'


# ---------------------------------------------------------------- CSS
CSS = """
:root{
  --teal:#32bfb6; --teal-lt:#94e1da; --ink:#2f302f; --sand:#dfc0aa;
  --teal-bg:#eafaf8; --teal-bg2:#f4fbfa; --sand-bg:#f7efe6;
  --line:#d9d9d6; --muted:#8a8f8d;
}
@page { size: A4; margin: 13mm 11mm 13mm 11mm; }
* { box-sizing: border-box; }
body { font-family:'Pretendard','Malgun Gothic',sans-serif; color:var(--ink); margin:0; font-size:13px; }
.page { page-break-after: always; }
.page:last-child { page-break-after: auto; }

.phead { display:flex; align-items:flex-end; justify-content:space-between;
  border-bottom:3px solid var(--teal); padding-bottom:6px; margin-bottom:11px; }
.ph-day { font-size:18px; font-weight:800; color:var(--teal); letter-spacing:.5px; }
.ph-title { text-align:center; }
.pt-ko { display:block; font-size:19px; font-weight:800; color:var(--ink); }
.pt-en { display:block; font-size:10px; color:var(--sand); letter-spacing:1.5px; text-transform:uppercase; font-weight:700; }
.ph-name { font-size:12px; color:var(--muted); }
.nline { display:inline-block; width:120px; border-bottom:1px solid var(--sand); margin-left:4px; }

/* 페이지 로고 (매 페이지 우하단 반복) */
.pagelogo { position:fixed; bottom:1mm; right:2mm; width:30mm; height:auto; opacity:.95; }
.guide { background:var(--teal-bg); border-left:4px solid var(--teal); padding:7px 11px; margin:0 0 12px;
  font-size:12.5px; color:var(--ink); border-radius:3px; }
.guide b { color:var(--teal); }

/* 단어표 */
table.wl { width:100%; border-collapse:collapse; }
table.wl th { background:var(--teal); color:#fff; font-size:12px; padding:8px 6px; font-weight:700; }
table.wl td { border:1px solid var(--line); padding:var(--wlp,7px) 7px; vertical-align:top; font-size:var(--wlf,12.5px); line-height:1.3; }
table.wl tr:nth-child(even) td { background:var(--teal-bg2); }
.c-no { text-align:center; color:var(--muted); width:26px; }
.c-eng { font-weight:800; color:var(--ink); width:82px; }
.c-mean { width:150px; }
.c-ex { color:#55605d; font-style:italic; }
.c-ant, .c-syn { width:80px; color:#6b6f6d; text-align:center; }

/* 쓰기 (뜻 먼저 -> 가로 쓰기줄) */
table.wr { width:100%; border-collapse:collapse; }
table.wr th { background:var(--teal); color:#fff; padding:8px; font-size:12px; font-weight:700; }
table.wr th.wh-no{ width:38px; } table.wr th.wh-eng{ width:120px; } table.wr th.wh-mean{ width:190px; }
table.wr td { border-bottom:1px solid var(--line); height:var(--wrh,38px); vertical-align:middle; }
.w-no { text-align:center; color:var(--muted); font-size:12px; }
.w-eng { font-weight:800; padding-left:12px; color:var(--teal); font-size:14px; border-left:1px dashed var(--teal-lt); }
.w-mean { padding-left:10px; color:var(--ink); font-size:var(--wmf,12.5px); line-height:1.25;
  border-left:1px dashed var(--teal-lt); }
.w-write { padding:0 10px; white-space:nowrap; border-left:1px dashed var(--teal-lt); }
.wl3 { display:inline-block; width:29%; border-bottom:1.4px solid var(--line); margin:0 1.5%; height:20px; }

/* 예문 빈칸 */
.bank { border:1.5px dashed var(--teal-lt); background:var(--teal-bg); border-radius:6px; padding:8px 10px;
  margin-bottom:11px; line-height:1.95; }
.chip { display:inline-block; background:#fff; border:1px solid var(--teal-lt); border-radius:13px;
  padding:2px 11px; margin:2px 3px; font-size:12px; font-weight:600; color:var(--ink); }
ol.fb { margin:0; padding-left:22px; }
ol.fb li { margin-bottom:var(--fbm,17px); line-height:var(--fbl,1.7); page-break-inside:avoid; break-inside:avoid; }
.fb-sent { font-size:var(--fbf,13px); }
.fb-mean { font-size:11px; color:var(--muted); margin-left:6px; }
.blank { display:inline-block; min-width:var(--fbb,120px); border-bottom:1.6px solid var(--ink); }
.ans { color:#c47a3d; font-weight:800; margin-left:4px; }

/* 복습 시험 */
table.rv { width:100%; border-collapse:collapse; }
table.rv td { border-bottom:1px solid var(--line); padding:var(--rvp,9px) 7px; font-size:var(--rvf,13px); }
.rv-no { width:28px; text-align:center; color:var(--muted); }
/* --rvq: --flat-review에서 문제가 한 줄에 들어가도록 넓힌 문제 칸 너비(기본은 기존 180px) */
.rv-eng { font-weight:800; width:var(--rvq,180px); color:var(--teal); }
.rv-mean { width:var(--rvq,180px); }
.rv-a { }
.wline { display:inline-block; width:90%; border-bottom:1px solid var(--line); }

/* 고르기 */
/* 고르기 항목: <li>는 Chrome 인쇄에서 break-inside가 무시되므로 블록 div로 만든다 */
.ch-list { margin:0; padding-left:4px; }
/* --chm/--chq/--cho/--chop: 항목이 많은 유닛에서 한 장에 넣기 위한 축소값(기본값 = 원래 크기) */
.ch-item { margin-bottom:var(--chm,4px); padding-left:16px; text-indent:-16px;
  page-break-inside:avoid; break-inside:avoid; }
.ch-item > * { text-indent:0; }
.ch-no { color:var(--muted); font-size:var(--chq,11.5px); }
.ch-q { font-size:var(--chq,11.5px); margin-bottom:1px; }
.ch-q b { color:var(--teal); }
.ch-mean { color:var(--muted); font-size:calc(var(--chq,11.5px) - 0.5px); font-weight:400; }
/* flex를 쓰면 Chrome 인쇄에서 break-inside:avoid가 무시돼 문제/선택지가 갈라짐 -> inline-block */
.ch-opts { display:block; page-break-inside:avoid; break-inside:avoid; }
.opt { display:inline-block; border:1px solid var(--line); border-radius:5px;
  padding:var(--chop,1px) 8px; font-size:var(--cho,11px); margin:0 4px 2px 0; }
.opt-ans { background:var(--sand-bg); border-color:var(--sand); color:#b06a2c; font-weight:800; }
.empty { color:var(--muted); }
.ch-sec { font-size:12.5px; font-weight:800; color:var(--teal); margin:7px 0 4px; }
.ch-sec:first-of-type { margin-top:2px; }
.ch-hint { font-size:11px; font-weight:400; color:var(--muted); }

/* 암기 노트 (딸기케이크식 4단 자가시험) */
table.mz { width:100%; border-collapse:collapse; table-layout:fixed; }
table.mz th { background:var(--teal); color:#fff; font-size:11px; font-weight:700; padding:6px 4px; }
.mz-th-sub { font-weight:400; font-size:8.5px; opacity:.9; }
table.mz td { border-bottom:1px solid var(--line); height:var(--mzh,26px); }
table.mz tr:nth-child(even) td { background:var(--teal-bg2); }
.mz-fold { border-left:1px dashed var(--teal-lt); }
.mz-no { text-align:center; color:var(--muted); font-size:11px; }
.mz-mark { text-align:center; color:#b4b2a9; }
.mz-oa { margin-top:10px; border:1.5px dashed var(--sand); background:var(--sand-bg); border-radius:6px; padding:8px 10px; }
.mz-oa-t { font-size:11.5px; font-weight:800; color:#b06a2c; margin-bottom:7px; }
.mz-oa-line { border-bottom:1px solid #d9c3ab; height:19px; margin-bottom:6px; }
.mz-oa-line:last-child { margin-bottom:0; }
"""

# --layout v2 전용 덧씌우기 (Bricks 4800).
# 지면 1장 = .page 1개로 높이를 A4 본문 상자에 고정해 로고를 지면 기준으로 배치한다.
# (A4 297mm - 상하 여백 26mm = 271mm. 반올림 오차로 빈 페이지가 생기지 않게 1mm 뺌)
CSS_V2 = """
/* --layout v2 와 --logo-top 이 공유하는 로고 배치용 규칙 */
.page { position:relative; height:270mm; padding-bottom:14mm; }
.page-hl { padding-bottom:0; }          /* 로고를 머리말로 올린 지면은 하단 여백 불필요 */
.pagelogo { position:absolute; }
.ph-logo { width:24mm; vertical-align:middle; margin-left:12px; }
.phead-lg .ph-name { display:flex; align-items:center; }
"""


# ---------------------------------------------------------------- 유닛 페이지 조립
def build_unit_pages(units, i, answer=False, merge_choice=False, merge_practice=False):
    """units: [(name, words)]. i: 0-based 현재 유닛 index."""
    name, words = units[i]
    base = (i + 1) * 1000
    pages = []
    pages.extend(page_wordlist(name, words))
    pages.extend(page_memorize(name, words))   # 단어목록 뒤 → '외우는 단계' 강제
    pages.extend(page_writing(name, words, seed=base + 1))
    # 누적 복습: N-2, N-1
    if i - 2 >= 0:
        tname, twords = units[i - 2]
        pages.append(page_review(name, tname, twords, seed=base + 200 + i, answer=answer))
    if i - 1 >= 0:
        tname, twords = units[i - 1]
        pages.append(page_review(name, tname, twords, seed=base + 300 + i, answer=answer))
    if merge_practice:  # 예문+동의+반의를 한 페이지로 (단어 아주 적은 교재)
        pages.append(page_practice_merged(name, words, seed=base, answer=answer))
    else:
        pages.append(page_fillblank(name, words, seed=base + 4, answer=answer))
        if merge_choice:  # 동의/반의만 한 페이지로 합쳐 종이 절약
            pages.extend(page_choice_merged(name, words, seed=base + 5, answer=answer))
        else:
            pages.extend(page_choice(name, words, "antonyms", seed=base + 5, answer=answer))
            pages.extend(page_choice(name, words, "synonyms", seed=base + 6, answer=answer))
    return pages


def font_face_css():
    """설치된 Pretendard(ttf)를 file:// 로 임베드해 헤드리스 Chrome에서 확실히 로드."""
    dirs = [
        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Windows\Fonts"),
        r"C:\Windows\Fonts",
    ]
    weights = {"Regular": 400, "Medium": 500, "SemiBold": 600, "Bold": 700, "ExtraBold": 800}
    rules = []
    for name, wt in weights.items():
        for d in dirs:
            fp = os.path.join(d, f"Pretendard-{name}.ttf")
            if os.path.exists(fp):
                url = "file:///" + fp.replace("\\", "/")
                rules.append(
                    "@font-face{font-family:'Pretendard';font-style:normal;"
                    f"font-weight:{wt};src:url('{url}') format('truetype');}}"
                )
                break
    return "\n".join(rules)


_LOGO_CACHE = None


def logo_datauri():
    """source/logo.png -> data URI. 없으면 빈 문자열."""
    global _LOGO_CACHE
    if _LOGO_CACHE is not None:
        return _LOGO_CACHE
    import base64
    here = os.path.dirname(os.path.abspath(__file__))
    fp = os.path.join(here, "source", "logo.png")
    if not os.path.exists(fp):
        _LOGO_CACHE = ""
    else:
        with open(fp, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        _LOGO_CACHE = "data:image/png;base64," + b64
    return _LOGO_CACHE


def build_html(book_name, units, day_from, day_to, answer=False, title_suffix="",
               merge_choice=False, merge_practice=False):
    body = []
    for i in range(day_from - 1, min(day_to, len(units))):
        body.extend(build_unit_pages(units, i, answer=answer,
                                     merge_choice=merge_choice, merge_practice=merge_practice))
    title = f"{book_name}{title_suffix} 워크북" + (" (정답)" if answer else "")
    # v1: body에 fixed 로고 하나가 전 지면에 반복. v2: 지면마다 page_foot()/머리말 로고.
    logo = logo_datauri()
    logo_html = "" if logo_on_top() or not logo else f'<img class="pagelogo" src="{logo}" alt="logo">'
    return f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<title>{esc(title)}</title><style>{font_face_css()}
{CSS}{CSS_V2 if logo_on_top() else ""}</style></head>
<body>{logo_html}{''.join(body)}</body></html>"""


# ---------------------------------------------------------------- PDF 변환
def find_chrome():
    cands = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    ]
    for c in cands:
        if os.path.exists(c):
            return c
    return None


def html_to_pdf(html_path, pdf_path):
    chrome = find_chrome()
    if not chrome:
        raise RuntimeError("Chrome/Edge를 찾을 수 없습니다.")
    import tempfile, shutil
    url = "file:///" + os.path.abspath(html_path).replace("\\", "/")
    out = os.path.abspath(pdf_path).replace("\\", "/")
    profile = tempfile.mkdtemp(prefix="vocawb_")  # 새 프로필 -> 실행 중인 Chrome과 독립
    try:
        r = subprocess.run([
            chrome, "--headless=new", "--disable-gpu", "--no-pdf-header-footer",
            f"--user-data-dir={profile}",
            f"--print-to-pdf={out}", url
        ], timeout=180, capture_output=True, encoding="utf-8", errors="replace")
        log = (r.stderr or "") + (r.stdout or "")
        if "written to file" not in log:
            raise RuntimeError(
                "PDF 생성 실패 (파일이 열려 있어 잠겨있을 수 있습니다 — 뷰어를 닫고 다시 시도하세요).\n"
                + log[-400:])
    finally:
        shutil.rmtree(profile, ignore_errors=True)
    stamp_page_numbers(out)


def stamp_page_numbers(pdf_path):
    """생성된 PDF의 각 페이지 하단 중앙에 페이지 번호를 찍는다(권마다 1부터).
    Chrome이 CSS @page 여백박스를 지원하지 않아 fitz로 후처리."""
    try:
        import fitz
    except ImportError:
        return  # PyMuPDF 없으면 번호 생략(치명적 아님)
    doc = fitz.open(pdf_path)
    for i, page in enumerate(doc, 1):
        r = page.rect
        txt = str(i)
        tw = fitz.get_text_length(txt, fontname="helv", fontsize=9)
        page.insert_text((r.width / 2 - tw / 2, r.height - 18),
                         txt, fontname="helv", fontsize=9, color=(0.5, 0.5, 0.5))
    doc.save(pdf_path, incremental=True, encryption=fitz.PDF_ENCRYPT_KEEP)
    doc.close()


# ---------------------------------------------------------------- 분권 범위 계산
def volume_ranges(total, n):
    """total개 유닛을 권당 n개로 분할. 마지막에 남는 자투리(<n)는 마지막 권에 합침.
    반환: [(from, to), ...] (1-based, 양끝 포함). n<=0 또는 n>=total 이면 단권."""
    if n <= 0 or n >= total:
        return [(1, total)]
    ranges = []
    start = 1
    while start + n - 1 <= total:
        ranges.append((start, start + n - 1))
        start += n
    if start <= total:  # 자투리 -> 마지막 권에 흡수
        s, _ = ranges[-1]
        ranges[-1] = (s, total)
    return ranges


def _emit(out, safe, book_name, units, a, b, vol_tag, title_suffix, answer,
          merge_choice=False, merge_practice=False):
    """한 범위(권)에 대해 학생용(+정답) HTML/PDF 생성."""
    h = build_html(book_name, units, a, b, answer=False, title_suffix=title_suffix,
                   merge_choice=merge_choice, merge_practice=merge_practice)
    hp = os.path.join(out, f"{safe}{vol_tag}_DAY{a}-{b}.html")
    with open(hp, "w", encoding="utf-8") as f:
        f.write(h)
    html_to_pdf(hp, hp[:-5] + ".pdf")
    print("생성:", hp[:-5] + ".pdf")
    if answer:
        ha = build_html(book_name, units, a, b, answer=True, title_suffix=title_suffix,
                        merge_choice=merge_choice, merge_practice=merge_practice)
        hap = os.path.join(out, f"{safe}{vol_tag}_DAY{a}-{b}_정답.html")
        with open(hap, "w", encoding="utf-8") as f:
            f.write(ha)
        html_to_pdf(hap, hap[:-5] + ".pdf")
        print("생성:", hap[:-5] + ".pdf")


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--from", dest="dfrom", type=int, default=1)
    ap.add_argument("--to", dest="dto", type=int, default=3)
    ap.add_argument("--out", default="output")
    ap.add_argument("--answer", action="store_true", help="정답 버전도 생성")
    ap.add_argument("--split", type=int, default=0,
                    help="권당 유닛 수로 분권 (예: --split 10). 0이면 단권(--from/--to 사용).")
    ap.add_argument("--merge-choice", dest="merge_choice", action="store_true",
                    help="동의어/반의어 고르기를 한 페이지로 합침 (동의/반의 적은 교재 종이 절약).")
    ap.add_argument("--merge-practice", dest="merge_practice", action="store_true",
                    help="예문 빈칸+동의어+반의어를 한 페이지로 합침 (단어 아주 적은 교재. --merge-choice 대체).")
    ap.add_argument("--fit-choice", dest="fit_choice", action="store_true",
                    help="--merge-choice 지면이 조금 넘칠 때 여백/글자를 줄여 한 장에 넣음 "
                         f"(최대 {CHOICE_FIT_MAX}칸까지. 그 이상은 그대로 다음 장으로).")
    ap.add_argument("--write-split", dest="write_split", type=int, default=0, metavar="N",
                    help="쓰기 연습을 N행씩 나눔(0=한 장). 예: 30단어를 --write-split 20 → 20+10 두 장.")
    ap.add_argument("--split-wordlist", dest="split_wordlist", type=int, default=0, metavar="N",
                    help="단어 목록을 N개씩 나눠 여러 장으로(제본용). 글자도 13.5px로 키운다.")
    ap.add_argument("--memo-split", dest="memo_split", type=int, default=0, metavar="N",
                    help="단어 암기 노트를 N행씩 나눔(0=한 장). 30단어에서 행이 눌려 못 쓸 때.")
    ap.add_argument("--balance-choice", dest="balance_choice", action="store_true",
                    help="고르기가 두 장으로 나뉠 때 앞 장을 꽉 채우지 않고 고르게 나눔.")
    ap.add_argument("--fit-wordlist", dest="fit_wordlist", action="store_true",
                    help="단어 목록이 넘칠 때 여백/글자를 더 줄여 한 유닛을 한 장에.")
    ap.add_argument("--flat-review", dest="flat_review", action="store_true",
                    help="누적 복습 시험의 문제가 두 줄로 접히지 않게 문제 칸을 넓힘.")
    ap.add_argument("--fill-review", dest="fill_review", action="store_true",
                    help="누적 복습 시험에서 아래로 남는 여백을 행 간격으로 돌려 지면을 채움.")
    ap.add_argument("--logo-top", dest="logo_top", action="store_true",
                    help="암기 노트·쓰기 연습은 로고를 머리말 오른쪽(학습한 날짜 뒤)으로 올리고 "
                         "나머지 지면은 우하단에. 하단이 비는 만큼 두 지면의 행을 키움.")
    ap.add_argument("--layout", choices=["v1", "v2"], default="v1",
                    help="v2: 지면별 로고 배치(암기노트·쓰기는 머리말) + 쓰기 연습 22행 분할. "
                         "현재 Bricks 4800 전용. 기본 v1은 기존 교재 지면 그대로.")
    args = ap.parse_args()

    global LAYOUT2, FIT_CHOICE, LOGO_TOP, WRITE_SPLIT, MEMO_SPLIT, BALANCE_CHOICE
    global WORDLIST_SPLIT
    global FIT_WORDLIST, FLAT_REVIEW, FILL_REVIEW
    LAYOUT2 = (args.layout == "v2")
    FIT_CHOICE = args.fit_choice
    LOGO_TOP = args.logo_top
    WRITE_SPLIT = args.write_split
    MEMO_SPLIT = args.memo_split
    WORDLIST_SPLIT = args.split_wordlist
    BALANCE_CHOICE = args.balance_choice
    FIT_WORDLIST = args.fit_wordlist
    FLAT_REVIEW = args.flat_review
    FILL_REVIEW = args.fill_review

    books = load_words(args.xlsx)
    os.makedirs(args.out, exist_ok=True)
    for book_name, units in books.items():
        safe = re.sub(r"[\\/:*?\"<>|]", "_", book_name)
        mc, mp = args.merge_choice, args.merge_practice
        if args.split and args.split > 0:
            ranges = volume_ranges(len(units), args.split)
            multi = len(ranges) > 1
            for vi, (a, b) in enumerate(ranges, 1):
                vol_tag = f"_{vi}권" if multi else ""
                title_suffix = f" {vi}권" if multi else ""
                _emit(args.out, safe, book_name, units, a, b, vol_tag, title_suffix, args.answer, mc, mp)
        else:
            _emit(args.out, safe, book_name, units, args.dfrom, args.dto, "", "", args.answer, mc, mp)


if __name__ == "__main__":
    main()
